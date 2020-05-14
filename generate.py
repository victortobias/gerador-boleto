from PyPDF2 import PdfFileWriter, PdfFileReader
import io
from reportlab.pdfgen import canvas
from reportlab.lib.pagesizes import letter
from reportlab.pdfbase import pdfmetrics
from reportlab.pdfbase.ttfonts import TTFont
from reportlab.graphics.barcode import code128
from barcode import ITF
from barcode.writer import ImageWriter
from PIL import Image
from math import ceil
import openpyxl

pdfmetrics.registerFont(TTFont('Calibri', 'calibri.ttf'))
pdfmetrics.registerFont(TTFont('Calibrib', 'calibrib.ttf'))
pdfmetrics.registerFont(TTFont('pixel', 'mp16sc.ttf'))

# ler a planilha de dados e criar um arquivo no zip no final
# wb = openpyxl.load_workbook("static"+os.path.sep+"upload"+os.path.sep+filename)
wb = openpyxl.load_workbook("base.xlsx")
ws = wb.worksheets[0]

maxRangeRow = ws.max_row
print(maxRangeRow)
for i in range(2, maxRangeRow):
    
    cpf = str(ws.cell(row=i, column=22).value)
    
    if((cpf != "CPF") or (cpf != None)):
        linha_digitavel = ws.cell(row=i, column=6).value
        carteira = ws.cell(row=i, column=18).value
        vencimento = ws.cell(row=i, column=12).value
        n_doc = ws.cell(row=i, column=11).value
        data_doc = ws.cell(row=i, column=16).value
        data_process = ws.cell(row=i, column=17).value
        agencia = ws.cell(row=i, column=8).value
        conta = ws.cell(row=i, column=9).value
        nosso_numero = ws.cell(row=i, column=10).value
        valor = ws.cell(row=i, column=4).value
        bene1 = str(ws.cell(row=i, column=21).value) + ' - CPF: '+ str(ws.cell(row=i, column=22).value)
        bene2 = str(ws.cell(row=i, column=24).value) + ' - ' + str(ws.cell(row=i, column=25).value)
        bene3 = str(ws.cell(row=i, column=27).value) + ' - ' + str(ws.cell(row=i, column=28).value) + ' - ' + str(ws.cell(row=i, column=26).value)
        bene1 = bene1.replace('None', '')
        bene2 = bene2.replace('None', '')
        bene3 = bene3.replace('None', '')
        cod_barras = ws.cell(row=i, column=20).value

        packet = io.BytesIO()
        # create a new PDF with Reportlab
        can = canvas.Canvas(packet, pagesize=letter)
        can.setFont('pixel', 12)
        can.drawString(250, 718, linha_digitavel) #linha digitavel superior
        can.drawString(250, 527, linha_digitavel) #linha digitavel inferior
        can.setFont('pixel', 9)
        can.drawString(174, 618, carteira) #Carteira superior
        can.drawString(140, 425, carteira) #Carteira inferior
        can.drawString(252, 638, "OUT") #Espécie Doc superior
        can.drawString(252, 445, "OUT") #Espécie Doc inferior
        can.drawString(318, 638, "0") #Aceite superior
        # can.drawString(318, 445, "0") #Aceite inferior
        can.drawString(493, 696, vencimento) #Vencimento superior
        can.drawString(493, 503, vencimento) #Vencimento inferior
        can.drawString(150, 638, n_doc) #Nº Doc superior
        can.drawString(150, 445, n_doc) #Nº Doc inferior
        can.drawString(60, 638, data_doc) #Data Doc superior
        can.drawString(60, 445, data_doc) #Data Doc inferior
        can.drawString(360, 638, data_process) #Data Proc superior
        can.drawString(360, 445, data_process) #Data Proc inferior
        can.drawString(470, 660, agencia+" / "+conta) #Agencia/Cod Beneficiario superior
        can.drawString(470, 468, agencia+" / "+conta) #Agencia/Cod Beneficiario inferior
        can.drawString(470, 638, nosso_numero) #Nosso Número superior
        can.drawString(470, 445, nosso_numero) #Nosso Número inferior
        can.drawString(470, 620, valor) #Valor do Documento superior
        can.drawString(470, 427, valor) #Valor do Documento inferior
        can.setFont('Calibrib', 8)
        can.drawString(106, 607, bene1.upper()) #Beneficiario 1ª nome linha superior
        can.drawString(106, 600, bene2.upper()) #Beneficiario 2ª endereço linha superior
        can.drawString(106, 593, bene3.upper()) #Beneficiario 3ª cep/cidade/uf linha superior
        can.drawString(106, 250, bene1.upper()) #Beneficiario 1ª nome linha inferior
        can.drawString(106, 243, bene2.upper()) #Beneficiario 2ª endereço linha inferior
        can.drawString(106, 236, bene3.upper()) #Beneficiario 3ª cep/cidade/uf linha inferior
        can.drawString(50, 400, '')#instruções 1
        can.drawString(50, 393, '')#instruções 2

        #Geração do codigo de barra modelo febraban
        itf = ITF(cod_barras, writer=ImageWriter())
        _ = itf.save('barcode')
        barcode = Image.open(r"barcode.png")
        box = (70, 2, 1000, 200)
        barcode = barcode.crop(box)#retira o numero gerado no codigo de barras
        barcode.save('barcode.png')

        can.drawImage('barcode.png', 40, 160, 400, 50, anchorAtXY=False) #codigo de barras


        can.save()

        #move to the beginning of the StringIO buffer
        packet.seek(0)
        new_pdf = PdfFileReader(packet)
        # read your existing PDF
        existing_pdf = PdfFileReader(open("branco.pdf", "rb"))
        output = PdfFileWriter()
        # add the "watermark" (which is the new pdf) on the existing page
        page = existing_pdf.getPage(0)
        page.mergePage(new_pdf.getPage(0))
        output.addPage(page)
        # finally, write "output" to a real file
        outputStream = open("b-"+cpf+".pdf", "wb")
        output.write(outputStream)
        outputStream.close()